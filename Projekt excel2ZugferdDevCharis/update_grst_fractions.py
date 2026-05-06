"""
Setzt in allen Positionsblättern der Musterdatei die Gebührensatz-Spalte (D)
auf Textbrüche (x/10 für Tab A/C, x/20 für Tab B) und passt die G-Formel an,
so dass der Bruch korrekt ausgewertet wird (LET/IFERROR statt D/10).
"""
import openpyxl
import os

EXCEL_PATH = r"Musterrechnungen\Excel2zugferd_Muster.xlsx"

# (Tab-Name, Zeile, Textbruch)
# Tab A/C: x/10  |  Tab B: D_zehntel*2/20  |  GrSt: Original "4/20" wiederherstellen
CHANGES = [
    ("Fibu",          14, "4/10"),
    ("Fibu",          15, "4,5/10"),   # 4,5 Zehntel (deutsches Dezimalkomma)
    ("Fibu",          16, "3/10"),
    ("JA",            14, "50/20"),    # 25 Zehntel = 50 Zwanzigstel (Tab B)
    ("JA",            16, "6/10"),
    ("GrSt",          14, "4/20"),     # Original wiederherstellen (Undo letzter Session)
    ("ESt Anlage V",  14, "10/10"),
    ("EÜR",           14, "56/20"),    # 28 Zehntel = 56 Zwanzigstel (Tab B)
    ("EÜR",           16, "6/10"),
    ("ESt",           14, "4/10"),
    ("ESt",           15, "8/10"),
    ("ESt",           16, "3,5/10"),   # 3,5 Zehntel
    ("ESt",           17, "2/10"),
]


def make_g_formula(r: int) -> str:
    """
    Neue G-Formel: wertet Textbruch in D aus (z.B. "4/10" → 0,4).
    Fallback auf numerischen D-Wert, falls D kein "/" enthält.
    Kein LET — kompatibel ab Excel 2010+.
    eval_D = IFERROR(VALUE(LEFT(D,FIND("/",D)-1))/VALUE(MID(D,FIND("/",D)+1,100)),D)
    """
    # Wiederholter Ausdruck: Bruch-Auswertung von D{r}
    eval_d = (
        f'IFERROR(VALUE(LEFT(D{r},FIND("/",D{r})-1))'
        f'/VALUE(MID(D{r},FIND("/",D{r})+1,100)),D{r})'
    )
    return (
        f'=IFERROR('
        f'IF(H{r}="",'
        f'{eval_d}*F{r},'
        f'{eval_d}*VLOOKUP(F{r},StBVV!$A:$D,MATCH(H{r},StBVV!$B$3:$D$3,0)+1,TRUE)),'
        f'{eval_d}*F{r})'
    )


def _recalculate_excel(path: str) -> None:
    abs_path = os.path.abspath(path)
    try:
        import win32com.client
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(abs_path)
        xl.Calculate()
        wb.Save()
        wb.Close(False)
        xl.Quit()
        print("  Formelcache aktualisiert.")
    except ImportError:
        print("  Hinweis: pywin32 nicht installiert — bitte Datei manuell öffnen")
    except Exception as e:
        print(f"  Hinweis: Formelcache konnte nicht automatisch aktualisiert werden: {e}")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    for sheet_name, row, frac_text in CHANGES:
        ws = wb[sheet_name]

        # D-Zelle: Textbruch setzen
        d_cell = ws.cell(row=row, column=4)
        d_cell.value = frac_text
        d_cell.data_type = "s"    # Explizit String
        d_cell.number_format = "@"  # Excel-Textformat

        # G-Zelle: neue Formel
        g_cell = ws.cell(row=row, column=7)
        g_cell.value = make_g_formula(row)

        print(f"  {sheet_name:15} Z{row}: D={frac_text!r:12}  G=LET(frac,...)")

    wb.save(EXCEL_PATH)
    print("Gespeichert.")

    print("Aktualisiere Formelcache via Excel ...")
    _recalculate_excel(EXCEL_PATH)
    print("Fertig!")


if __name__ == "__main__":
    main()
