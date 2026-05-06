"""
Aktualisiert das StBVV-Tabellenblatt in Excel2zugferd_Muster.xlsx
mit den offiziellen Werten der 5. Änderungsverordnung
(BGBl. 2025 I Nr. 105, gültig ab 01.07.2025).

Struktur des Sheets (rückwärtskompatibel mit SVERWEIS-Formeln):
  Zeile 3 : Headers — A3="Gegenstandswert ab €", B3="A", C3="B", D3="C"
  Zeile 5+ : Daten — LB | Tabelle-A-Wert | Tabelle-B-Wert | Tabelle-C-Wert

Formel in Positionsblättern (unverändert):
  =IFERROR(IF(H="",D*F,D/10*VLOOKUP(F,StBVV!$A:$D,MATCH(H,StBVV!$B$3:$D$3,0)+1,TRUE)),D*F)
"""

import bisect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"Musterrechnungen\Excel2zugferd_Muster.xlsx"

# ─── Offizielle Tabelle A (Beratungstabelle) ab 01.07.2025 ──────────────────
# Format: (Gegenstandswert BIS €, volle Gebühr 10/10 €)
TABLE_A_BIS = [
    (300, 31), (600, 56), (900, 81), (1200, 106), (1500, 130),
    (2000, 166), (2500, 200), (3000, 235), (3500, 270), (4000, 305),
    (4500, 340), (5000, 375), (6000, 422), (7000, 467), (8000, 514),
    (9000, 560), (10000, 605), (13000, 655), (16000, 705), (19000, 755),
    (22000, 805), (25000, 854), (30000, 946), (35000, 1036), (40000, 1125),
    (45000, 1215), (50000, 1304), (65000, 1399), (80000, 1496),
    (95000, 1592), (110000, 1689), (125000, 1784), (140000, 1879),
    (155000, 1976), (170000, 2071), (185000, 2168), (200000, 2264),
    (230000, 2412), (260000, 2559), (290000, 2705), (320000, 2859),
    (350000, 2926), (380000, 2990), (410000, 3055), (440000, 3115),
    (470000, 3175), (500000, 3234), (550000, 3320), (600000, 3404),
]

# ─── Offizielle Tabelle B (Abschlusstabelle) ab 01.07.2025 ──────────────────
TABLE_B_BIS = [
    (3000, 49), (3500, 57), (4000, 68), (4500, 76), (5000, 86),
    (6000, 96), (7000, 105), (8000, 116), (9000, 121), (10000, 127),
    (12500, 134), (15000, 151), (17500, 166), (20000, 178), (22500, 191),
    (25000, 201), (37500, 215), (50000, 263), (62500, 303), (75000, 338),
    (87500, 353), (100000, 369), (125000, 423), (150000, 471),
    (175000, 512), (200000, 548), (225000, 582), (250000, 613),
    (300000, 641), (350000, 696), (400000, 746), (450000, 791),
    (500000, 832), (625000, 871), (750000, 968), (875000, 1050),
    (1000000, 1126), (1250000, 1194), (1500000, 1324), (1750000, 1438),
    (2000000, 1542),
]

# ─── Offizielle Tabelle C (Buchführungstabelle) ab 01.07.2025 ────────────────
TABLE_C_BIS = [
    (15000, 72), (17500, 80), (20000, 88), (22500, 93), (25000, 101),
    (30000, 108), (35000, 117), (40000, 122), (45000, 129), (50000, 138),
    (62500, 145), (75000, 158), (87500, 174), (100000, 188), (125000, 209),
    (150000, 230), (200000, 275), (250000, 317), (300000, 359),
    (350000, 404), (400000, 441), (450000, 475), (500000, 512),
    # Übersteigungsregel (Anlage 3 StBVV 2025):
    # je angefangene 50.000 € über 500.000 € +36 €, bis 5.000.000 €
    *[(500000 + i * 50000, 512 + i * 36) for i in range(1, 91)],
]


def table_value_for_lb(lb: int, table_bis: list[tuple]) -> int | None:
    """Gibt den Gebührenwert aus 'table_bis' zurück, der für die
    Untergrenze 'lb' gilt. 'lb' ist der erste Wert eines Tiers.
    Rückgabe: Gebühr falls der LB in den Tabellen-Bereich fällt, sonst None."""
    # Das passende Tier ist dasjenige mit dem kleinsten BIS-Wert >= lb.
    # (Denn: LB N gehört zum Tier "bis BIS_N = Wert", wobei BIS_N >= lb.)
    bis_vals = [b for b, _ in table_bis]
    idx = bisect.bisect_left(bis_vals, lb)
    if idx < len(table_bis):
        return table_bis[idx][1]
    return None   # über dem Tabellenmaximum


def build_combined_table() -> list[tuple]:
    """Baut die kombinierte LB-Tabelle (SVERWEIS-kompatibel).
    Gibt Liste von (lb, val_A, val_B, val_C) zurück."""
    # Alle einzigartigen BIS-Werte sammeln
    all_bis = sorted({b for b, _ in TABLE_A_BIS + TABLE_B_BIS + TABLE_C_BIS})

    # LBs: erster LB = 1, danach BIS[i-1] + 1
    lbs = [1] + [b + 1 for b in all_bis[:-1]]

    rows = []
    for lb in lbs:
        val_a = table_value_for_lb(lb, TABLE_A_BIS)
        val_b = table_value_for_lb(lb, TABLE_B_BIS)
        val_c = table_value_for_lb(lb, TABLE_C_BIS)
        rows.append((lb, val_a, val_b, val_c))
    return rows


def make_thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def make_medium_bottom():
    t = Side(style="thin")
    m = Side(style="medium")
    return Border(left=t, right=t, top=t, bottom=m)

def write_stbvv_sheet(ws):
    # ── Alles löschen ────────────────────────────────────────────────────────
    # Merge-Bereiche erst aufheben, dann Zeilen löschen
    ws.merged_cells.ranges.clear()
    ws.delete_rows(1, ws.max_row + 10)

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 3

    FILL_HEADER = PatternFill("solid", fgColor="2E5596")   # dunkelblau
    FILL_SUB    = PatternFill("solid", fgColor="BDD7EE")   # hellblau
    FILL_ODD    = PatternFill("solid", fgColor="F2F7FC")
    FILL_EVEN   = PatternFill("solid", fgColor="FFFFFF")
    FILL_DATE   = PatternFill("solid", fgColor="DEEAF1")
    FILL_LEGEND = PatternFill("solid", fgColor="FFF2CC")

    WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=11)
    DARK_BOLD   = Font(bold=True, size=10)
    ITALIC_GREY = Font(italic=True, size=9, color="595959")

    border = make_thin()
    border_hdr = make_medium_bottom()

    # ── Zeile 1: Haupttitel ──────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "StBVV — Gebührentabellen A, B und C (Anlage 1–3)"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Zeile 2: Datum / Rechtsgrundlage ─────────────────────────────────────
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = (
        "Gültig ab 01.07.2025 · "
        "5. Verordnung zur Änderung der StBVV v. 31.03.2025 (BGBl. 2025 I Nr. 105)"
    )
    c.font = ITALIC_GREY
    c.fill = FILL_DATE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    # ── Zeile 3: SVERWEIS-Spaltenköpfe (PFLICHT – MATCH sucht nach "A","B","C")
    headers = [("A3", "Gegenstandswert ab €"), ("B3", "A"), ("C3", "B"), ("D3", "C")]
    for addr, val in headers:
        c = ws[addr]
        c.value = val
        c.font = WHITE_BOLD
        c.fill = FILL_HEADER
        c.border = border_hdr
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].alignment = Alignment(horizontal="left", indent=1, vertical="center")
    ws.row_dimensions[3].height = 18

    # ── Zeile 4: Erläuterungen (kursiv) ──────────────────────────────────────
    sub_labels = [
        ("A4", "(Untergrenze des Tiers für SVERWEIS)"),
        ("B4", "Beratungstabelle"),
        ("C4", "Abschlusstabelle"),
        ("D4", "Buchführung (monatl.)"),
    ]
    for addr, val in sub_labels:
        c = ws[addr]
        c.value = val
        c.font = Font(italic=True, size=8, color="595959")
        c.fill = FILL_SUB
        c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    # ── Legende: Verwendungshinweis ───────────────────────────────────────────
    # (Wir fügen sie NACH den Daten ein — wird am Ende gesetzt)

    # ── Daten ab Zeile 5 ─────────────────────────────────────────────────────
    combined = build_combined_table()
    DATA_START = 5
    for i, (lb, va, vb, vc) in enumerate(combined):
        r = DATA_START + i
        fill = FILL_ODD if i % 2 == 0 else FILL_EVEN

        c_lb = ws.cell(r, 1, value=lb)
        c_lb.number_format = '#,##0'
        c_lb.alignment = Alignment(horizontal="right")

        c_a = ws.cell(r, 2, value=va)
        c_b = ws.cell(r, 3, value=vb)
        c_c = ws.cell(r, 4, value=vc)

        for c in (c_lb, c_a, c_b, c_c):
            c.fill = fill
            c.border = border
            c.alignment = Alignment(horizontal="right")
            if c.value is not None and c.column > 1:
                c.number_format = '#,##0'

    # ── Fußzeile: Legende ─────────────────────────────────────────────────────
    legend_row = DATA_START + len(combined) + 1
    ws.merge_cells(
        start_row=legend_row, start_column=1,
        end_row=legend_row, end_column=4
    )
    c = ws.cell(legend_row, 1)
    c.value = (
        "Spalte H im Positionsbereich: A = Beratung/Steuererklärung, "
        "B = Abschluss/EÜR, C = Buchführung (§ 33 StBVV) · "
        "Werte > Tabellenmax: individuelle Berechnung erforderlich"
    )
    c.font = ITALIC_GREY
    c.fill = FILL_LEGEND
    c.alignment = Alignment(horizontal="left", wrap_text=True, indent=1)
    ws.row_dimensions[legend_row].height = 28

    print(f"  Kombinierte Tabelle: {len(combined)} Zeilen (Untergrenzen)")
    print(f"  Daten von Zeile {DATA_START} bis {DATA_START + len(combined) - 1}")


def main():
    print(f"Öffne {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if "StBVV" not in wb.sheetnames:
        print(f"FEHLER: Blatt 'StBVV' nicht gefunden! Vorhandene: {wb.sheetnames}")
        return

    ws = wb["StBVV"]
    print("Schreibe aktualisierte StBVV-Tabellen (gültig ab 01.07.2025) ...")
    write_stbvv_sheet(ws)

    print("Speichere ...")
    wb.save(EXCEL_PATH)
    print("Aktualisiere Formelcache via Excel ...")
    _recalculate_excel(EXCEL_PATH)
    print("Fertig!")


def _recalculate_excel(path: str) -> None:
    """Öffnet die Excel-Datei via COM, berechnet alle Formeln und speichert."""
    import os
    abs_path = os.path.abspath(path)
    try:
        import win32com.client
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(abs_path)
        xl.Calculate()
        wb.Save()
        wb.Close(False)
        xl.Quit()
    except ImportError:
        print("  Hinweis: pywin32 nicht installiert — bitte Datei manuell in Excel öffnen und speichern.")
    except Exception as e:
        print(f"  Hinweis: Formelcache konnte nicht automatisch aktualisiert werden: {e}")
        print("  Bitte Datei manuell in Excel öffnen und speichern.")


if __name__ == "__main__":
    main()
