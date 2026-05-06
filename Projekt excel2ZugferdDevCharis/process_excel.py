import openpyxl
import os

file1 = r"c:\Users\Charis\Projekte\excel2zugferd\Musterrechnungen\10117 Klaus Walter.xlsx"
file2 = r"c:\Users\Charis\Projekte\excel2zugferd\Musterrechnungen\Excel2zugferd_Muster.xlsx"

print(f"--- File: {os.path.basename(file1)} ---")
wb1 = openpyxl.load_workbook(file1, data_only=True)
print(f"Sheet names: {wb1.sheetnames}")

target_sheet = "GrSt 2022"
if target_sheet in wb1.sheetnames:
    ws = wb1[target_sheet]
    print(f"\n--- Sheet: {target_sheet} ---")
    # For large sheets, we might want to cap the rows, but the request says all rows.
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"Row {i}: {row}")
else:
    print(f"\nSheet {target_sheet} not found.")

print(f"\n--- File: {os.path.basename(file2)} ---")
wb2 = openpyxl.load_workbook(file2)
print(f"Sheet names: {wb2.sheetnames}")
